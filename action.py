import pyautogui
import re
import time
import os
import threading
import pywhatkit
import csv
from io import StringIO
from openpyxl import Workbook, load_workbook

# Import AppOpener functions
from AppOpener import open as open_app
from AppOpener import close as close_app
from AppOpener import give_appnames

class ExcelCommandHandler:
    def __init__(self):
        self.default_extension = ".xlsx"

    def _normalize_path(self, file_name):
        clean_name = file_name.strip().strip('"').strip("'")
        if not clean_name.lower().endswith(self.default_extension):
            clean_name += self.default_extension

        is_windows_absolute = re.match(r"^[a-zA-Z]:\\", clean_name) is not None
        if not os.path.isabs(clean_name) and not is_windows_absolute:
            clean_name = os.path.abspath(clean_name)

        parent = os.path.dirname(clean_name)
        if parent:
            os.makedirs(parent, exist_ok=True)
        return clean_name

    def _load_or_create_workbook(self, path):
        if os.path.exists(path):
            return load_workbook(path)
        return Workbook()

    def _pick_sheet(self, workbook, sheet_name):
        if not sheet_name:
            return workbook.active
        target_sheet = sheet_name.strip().strip('"').strip("'")
        if target_sheet in workbook.sheetnames:
            return workbook[target_sheet]
        return workbook.create_sheet(target_sheet)

    def _to_value(self, raw_value):
        value = raw_value.strip()
        if (value.startswith('"') and value.endswith('"')) or (value.startswith("'") and value.endswith("'")):
            return value[1:-1]
        if value.lower() == "true":
            return True
        if value.lower() == "false":
            return False
        if re.fullmatch(r"-?\d+", value):
            return int(value)
        if re.fullmatch(r"-?\d+\.\d+", value):
            return float(value)
        return value

    def handle(self, text):
        open_match = re.fullmatch(r"excel\s+(?:create|open)\s+(.+)", text, flags=re.IGNORECASE)
        if open_match:
            file_path = self._normalize_path(open_match.group(1))
            workbook = self._load_or_create_workbook(file_path)
            workbook.save(file_path)
            print(f"[ACTION][EXCEL] Ready: {file_path}")
            return True

        set_match = re.fullmatch(
            r"excel\s+set\s+([a-zA-Z]+\d+)\s+to\s+(.+?)\s+in\s+(.+?)(?:\s+sheet\s+(.+))?",
            text,
            flags=re.IGNORECASE,
        )
        if set_match:
            cell, value_text, file_name, sheet_name = set_match.groups()
            file_path = self._normalize_path(file_name)
            workbook = self._load_or_create_workbook(file_path)
            sheet = self._pick_sheet(workbook, sheet_name)
            sheet[cell.upper()] = self._to_value(value_text)
            workbook.save(file_path)
            print(f"[ACTION][EXCEL] {cell.upper()} updated in {file_path}")
            return True

        get_match = re.fullmatch(
            r"excel\s+get\s+([a-zA-Z]+\d+)\s+in\s+(.+?)(?:\s+sheet\s+(.+))?",
            text,
            flags=re.IGNORECASE,
        )
        if get_match:
            cell, file_name, sheet_name = get_match.groups()
            file_path = self._normalize_path(file_name)
            workbook = self._load_or_create_workbook(file_path)
            sheet = self._pick_sheet(workbook, sheet_name)
            value = sheet[cell.upper()].value
            workbook.save(file_path)
            print(f"[ACTION][EXCEL] {cell.upper()} = {value}")
            return True

        row_match = re.fullmatch(
            r"excel\s+add\s+row\s+(.+?)\s+in\s+(.+?)(?:\s+sheet\s+(.+))?",
            text,
            flags=re.IGNORECASE,
        )
        if row_match:
            row_values_text, file_name, sheet_name = row_match.groups()
            file_path = self._normalize_path(file_name)
            workbook = self._load_or_create_workbook(file_path)
            sheet = self._pick_sheet(workbook, sheet_name)
            parsed_row = next(csv.reader(StringIO(row_values_text), skipinitialspace=True))
            sheet.append([self._to_value(item) for item in parsed_row])
            workbook.save(file_path)
            print(f"[ACTION][EXCEL] Row added in {file_path}")
            return True

        sum_match = re.fullmatch(
            r"excel\s+sum\s+([a-zA-Z]+\d+:[a-zA-Z]+\d+)\s+in\s+(.+?)\s+to\s+([a-zA-Z]+\d+)(?:\s+sheet\s+(.+))?",
            text,
            flags=re.IGNORECASE,
        )
        if sum_match:
            source_range, file_name, target_cell, sheet_name = sum_match.groups()
            file_path = self._normalize_path(file_name)
            workbook = self._load_or_create_workbook(file_path)
            sheet = self._pick_sheet(workbook, sheet_name)
            total = 0
            for row in sheet[source_range.upper()]:
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        total += cell.value
                    elif isinstance(cell.value, str) and re.fullmatch(r"-?\d+(\.\d+)?", cell.value.strip()):
                        total += float(cell.value)
            sheet[target_cell.upper()] = total
            workbook.save(file_path)
            print(f"[ACTION][EXCEL] Sum {source_range.upper()} -> {target_cell.upper()} in {file_path}")
            return True

        formula_match = re.fullmatch(
            r"excel\s+formula\s+([a-zA-Z]+\d+)\s*=\s*(.+?)\s+in\s+(.+?)(?:\s+sheet\s+(.+))?",
            text,
            flags=re.IGNORECASE,
        )
        if formula_match:
            target_cell, formula_text, file_name, sheet_name = formula_match.groups()
            file_path = self._normalize_path(file_name)
            workbook = self._load_or_create_workbook(file_path)
            sheet = self._pick_sheet(workbook, sheet_name)
            cleaned_formula = formula_text.strip()
            if not cleaned_formula.startswith("="):
                cleaned_formula = "=" + cleaned_formula
            sheet[target_cell.upper()] = cleaned_formula
            workbook.save(file_path)
            print(f"[ACTION][EXCEL] Formula set in {target_cell.upper()} for {file_path}")
            return True

        if re.fullmatch(r"excel\s+help", text, flags=re.IGNORECASE):
            print("[ACTION][EXCEL] Commands:")
            print("- excel create <file>")
            print("- excel set <cell> to <value> in <file> [sheet <sheet_name>]")
            print("- excel get <cell> in <file> [sheet <sheet_name>]")
            print("- excel add row <comma-separated values> in <file> [sheet <sheet_name>]")
            print("- excel sum <A1:B10> in <file> to <cell> [sheet <sheet_name>]")
            print("- excel formula <cell> = <formula> in <file> [sheet <sheet_name>]")
            return True

        return False

class ActionHandler:
    def __init__(self):
        # 1. CUSTOM APPS / GAMES
        # Add games or portable apps here that the scanner misses.
        # Use double backslashes \\ for paths.
        self.custom_apps = {
            "genshin": r"C:\Program Files\Genshin Impact\Genshin Impact Game\GenshinImpact.exe",
            "minecraft": r"C:\XboxGames\Minecraft Launcher\Content\Minecraft.exe",
            "steam": r"C:\Program Files (x86)\Steam\steam.exe",
            "obs": r"C:\Program Files\obs-studio\bin\64bit\obs64.exe"
        }
        self.excel = ExcelCommandHandler()

    def execute(self, text):
        if not text: return
        raw_text = text.strip()
        text = raw_text.lower()

        # =========================================================
        # EXCEL COMMANDS
        # =========================================================
        if text.startswith("excel ") and self.excel.handle(raw_text):
            return

        # =========================================================
        # 0. SPECIAL COMMAND: UPDATE APP LIST
        # =========================================================
        if "scan apps" in text or "update apps" in text:
            print("[ACTION] Scanning for new apps...")
            # Run this in a thread so it doesn't freeze MARIE
            threading.Thread(target=give_appnames, daemon=True).start()
            return

        # =========================================================
        # 1. YOUTUBE (Play Video)
        # =========================================================
        if text.startswith("play "):
            video_topic = text.replace("play ", "").replace("please", "").strip()
            print(f"[ACTION] Playing on YouTube: {video_topic}")
            try:
                pywhatkit.playonyt(video_topic)
            except Exception as e:
                print(f"[ERROR] YouTube failed: {e}")
            return

        # =========================================================
        # 2. NOTEPAD (Write text)
        # =========================================================
        write_triggers = ["write ", "note ", "type ", "take a note "]
        triggered_word = next((w for w in write_triggers if text.startswith(w)), None)

        if triggered_word:
            content = text.replace(triggered_word, "").strip()
            print(f"[ACTION] Writing to Notepad: {content}")
            os.system("start notepad") 
            time.sleep(1.0) # Wait for it to open
            pyautogui.write(content, interval=0.05)
            return

        # =========================================================
        # 3. SYSTEM CONTROLS
        # =========================================================
        if "volume up" in text:
            pyautogui.press('volumeup')
            return
        elif "volume down" in text:
            pyautogui.press('volumedown')
            return
        elif "mute" in text or "unmute" in text:
            pyautogui.press('volumemute')
            return

        # =========================================================
        # 4. OPEN APPS (Custom + General)
        # =========================================================
        if text.startswith("open "):
            raw_name = text.replace("open ", "").strip()
            app_name = raw_name.replace("please", "").replace("now", "").strip()
            app_name = re.sub(r'[^\w\s]', '', app_name)

            print(f"[ACTION] Opening: '{app_name}'")

            # A. Check Custom List first (Games/Portable)
            for key, path in self.custom_apps.items():
                if key in app_name:
                    print(f"[ACTION] Found custom path for {key}")
                    try:
                        os.startfile(path)
                    except Exception as e:
                        print(f"[ERROR] Custom path failed: {e}")
                    return

            # B. Check General List (AppOpener)
            try:
                open_app(app_name, match_closest=True, output=False, throw_error=True)
            except:
                # C. Last Resort: Windows Start
                try:
                    os.system(f"start {app_name}")
                except:
                    print(f"[ERROR] Could not open '{app_name}'")
            return

        # =========================================================
        # 5. CLOSE APPS
        # =========================================================
        if text.startswith("close "):
            app_name = text.replace("close ", "").replace("please", "").strip()
            app_name = re.sub(r'[^\w\s]', '', app_name)
            
            try:
                close_app(app_name, match_closest=True, output=False, throw_error=True)
            except:
                print(f"[ERROR] Could not close '{app_name}'")
            return
