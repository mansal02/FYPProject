import csv
import hashlib
import json
import os
import threading
import time
import re
from pathlib import Path

from aiassistant.infra.config.app_config import CONFIG

try:
    import PyPDF2
except Exception:  # pragma: no cover - optional dependency
    PyPDF2 = None

try:
    import docx
except Exception:  # pragma: no cover - optional dependency
    docx = None

try:
    import openpyxl
except Exception:  # pragma: no cover - optional dependency
    openpyxl = None


def _normalize_extensions(items):
    normalized = []
    for item in items or []:
        ext = str(item or "").strip().lower()
        if not ext:
            continue
        if not ext.startswith("."):
            ext = "." + ext
        normalized.append(ext)
    return sorted(set(normalized))


def _normalize_skip_dirs(items):
    normalized = []
    for item in items or []:
        raw = str(item or "").strip()
        if not raw:
            continue
        normalized.append(raw.replace("\\", "/").lower())
    return sorted(set(normalized))


def _default_training_config():
    training_cfg = CONFIG.get("training", {})
    paths_cfg = CONFIG.get("paths", {})
    return {
        "train_root": str(paths_cfg.get("train_root", "D:/Train")),
        "response_dir": str(paths_cfg.get("train_response_dir", "D:/Train/response")),
        "idle_min_sec": int(training_cfg.get("idle_min_sec", 120)),
        "idle_max_sec": int(training_cfg.get("idle_max_sec", 300)),
        "max_files_per_cycle": int(training_cfg.get("max_files_per_cycle", 24)),
        "sleep_between_files_sec": float(training_cfg.get("sleep_between_files_sec", 0.05)),
        "include_extensions": _normalize_extensions(training_cfg.get("include_extensions", [])),
        "skip_dirs": _normalize_skip_dirs(training_cfg.get("skip_dirs", [])),
    }


def _tokenize_words(text: str) -> list[str]:
    return re.findall(r"[a-zA-Z][a-zA-Z\-']{1,}", text.lower())


def _build_style_profile_from_texts(texts: list[str]) -> dict:
    if not texts:
        return {}

    combined = "\n".join(texts)
    words = _tokenize_words(combined)
    if not words:
        return {}

    stop_words = {
        "the",
        "and",
        "to",
        "of",
        "in",
        "a",
        "is",
        "for",
        "on",
        "with",
        "that",
        "this",
        "it",
        "as",
        "are",
        "be",
        "by",
        "an",
        "or",
        "from",
        "at",
        "was",
        "were",
    }
    freq = {}
    for word in words:
        if word in stop_words:
            continue
        freq[word] = freq.get(word, 0) + 1

    top_words = sorted(freq.items(), key=lambda item: item[1], reverse=True)[:20]
    top_vocab = [word for word, _ in top_words]

    sentences = re.split(r"[.!?]+", combined)
    sentence_lengths = [len(_tokenize_words(s)) for s in sentences if s.strip()]
    avg_sentence = int(sum(sentence_lengths) / max(1, len(sentence_lengths)))

    contractions = sum(1 for w in words if "'" in w)
    contraction_ratio = contractions / float(len(words))

    style_summary = {
        "avg_sentence_len": avg_sentence,
        "top_vocab": top_vocab,
        "contraction_ratio": round(contraction_ratio, 3),
    }

    formal = (
        "Formal baseline: Use clear, structured sentences; avoid slang; "
        "prefer precise wording. Incorporate these keywords when relevant: "
        + ", ".join(top_vocab[:10])
    )
    casual = (
        "Casual baseline: Use friendly, concise sentences; contractions are OK; "
        "keep tone approachable. Incorporate these keywords when relevant: "
        + ", ".join(top_vocab[:10])
    )

    return {
        "summary": style_summary,
        "formal": formal.strip(),
        "casual": casual.strip(),
    }


class DocumentIndexer:
    def __init__(
        self,
        db,
        roots,
        response_dir,
        include_extensions,
        skip_dirs,
        status_cb=None,
    ):
        self.db = db
        self.roots = [Path(str(root)).expanduser() for root in (roots or [])]
        self.response_dir = Path(str(response_dir or "")).expanduser()
        self.include_extensions = _normalize_extensions(include_extensions)
        self.skip_dirs = _normalize_skip_dirs(skip_dirs)
        self.status_cb = status_cb
        self._stop_event = threading.Event()
        self._log_path = None

    def ensure_response_dir(self):
        if not self.response_dir:
            return ""
        try:
            self.response_dir.mkdir(parents=True, exist_ok=True)
            self._log_path = self.response_dir / "mirror_index_log.jsonl"
            return str(self.response_dir)
        except Exception:
            return ""

    def stop(self):
        self._stop_event.set()

    def reset_stop(self):
        self._stop_event.clear()

    def _emit_status(self, message):
        if self.status_cb:
            try:
                self.status_cb(message)
            except Exception:
                pass

    def _should_skip_dir(self, path_obj: Path) -> bool:
        try:
            raw = str(path_obj.resolve()).replace("\\", "/").lower()
        except Exception:
            raw = str(path_obj).replace("\\", "/").lower()
        if any(raw.startswith(prefix) for prefix in self.skip_dirs if ":/" in prefix):
            return True
        for part in path_obj.parts:
            part_norm = str(part).lower()
            if part_norm.startswith("."):
                return True
            if part_norm in {".git", "__pycache__"}:
                return True
            if part_norm.replace("\\", "/").lower() in self.skip_dirs:
                return True
        return False

    def _iter_files_dfs(self):
        stack = []
        for root in self.roots:
            if root and root.exists():
                stack.append(root)

        while stack:
            current = stack.pop()
            if self._stop_event.is_set():
                return
            try:
                if current.is_dir():
                    if self._should_skip_dir(current):
                        continue
                    entries = []
                    try:
                        entries = list(current.iterdir())
                    except Exception:
                        entries = []
                    for entry in reversed(entries):
                        stack.append(entry)
                elif current.is_file():
                    if self.include_extensions:
                        if current.suffix.lower() not in self.include_extensions:
                            continue
                    yield current
            except Exception:
                continue

    def _hash_file(self, path_obj: Path):
        try:
            hasher = hashlib.sha256()
            with path_obj.open("rb") as handle:
                for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                    if not chunk:
                        break
                    hasher.update(chunk)
            return hasher.hexdigest()
        except Exception:
            return ""

    def _read_text_file(self, path_obj: Path):
        try:
            return path_obj.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            return ""

    def _read_pdf(self, path_obj: Path):
        if PyPDF2 is None:
            return ""
        try:
            reader = PyPDF2.PdfReader(str(path_obj))
            pages = []
            for page in reader.pages:
                text = page.extract_text() or ""
                pages.append(text)
            return "\n".join(pages)
        except Exception:
            return ""

    def _read_docx(self, path_obj: Path):
        if docx is None:
            return ""
        try:
            document = docx.Document(str(path_obj))
            paragraphs = [p.text for p in document.paragraphs if p.text]
            return "\n".join(paragraphs)
        except Exception:
            return ""

    def _read_csv(self, path_obj: Path):
        try:
            rows = []
            with path_obj.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
                reader = csv.reader(handle)
                for row in reader:
                    rows.append("\t".join(str(cell) for cell in row))
            return "\n".join(rows)
        except Exception:
            return ""

    def _read_xlsx(self, path_obj: Path):
        if openpyxl is None:
            return ""
        try:
            wb = openpyxl.load_workbook(str(path_obj), read_only=True, data_only=True)
            output = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                output.append(f"[Sheet] {sheet_name}")
                for row in sheet.iter_rows(values_only=True):
                    output.append("\t".join("" if cell is None else str(cell) for cell in row))
            return "\n".join(output)
        except Exception:
            return ""

    def _extract_text(self, path_obj: Path):
        suffix = path_obj.suffix.lower()
        if suffix in {".txt", ".log", ".md"}:
            return self._read_text_file(path_obj)
        if suffix == ".pdf":
            return self._read_pdf(path_obj)
        if suffix == ".docx":
            return self._read_docx(path_obj)
        if suffix == ".csv":
            return self._read_csv(path_obj)
        if suffix == ".xlsx":
            return self._read_xlsx(path_obj)
        return self._read_text_file(path_obj)

    def build_style_profile(self, style_root: Path, max_chars_per_file: int = 24000):
        texts = []
        for path_obj in self._iter_files_dfs():
            if self._stop_event.is_set():
                break
            try:
                if not str(path_obj).replace("\\", "/").lower().startswith(
                    str(style_root).replace("\\", "/").lower()
                ):
                    continue
            except Exception:
                continue
            raw = self._extract_text(path_obj)
            if not raw:
                continue
            texts.append(raw[:max_chars_per_file])

        profile = _build_style_profile_from_texts(texts)
        if profile and hasattr(self.db, "save_style_profile"):
            self.db.save_style_profile("train_root", profile)

        if self.response_dir:
            try:
                out_path = self.response_dir / "style_profile.json"
                out_path.write_text(json.dumps(profile, indent=2, ensure_ascii=True), encoding="utf-8")
            except Exception:
                pass
        return profile

    def _log_index_event(self, payload):
        if not self._log_path:
            return
        try:
            with self._log_path.open("a", encoding="utf-8") as handle:
                handle.write(json.dumps(payload, ensure_ascii=True) + "\n")
        except Exception:
            pass

    def run_index_cycle(self, max_files=24, sleep_sec=0.05):
        self.ensure_response_dir()
        self.reset_stop()
        processed = 0

        for path_obj in self._iter_files_dfs():
            if self._stop_event.is_set():
                break

            file_hash = self._hash_file(path_obj)
            if not file_hash:
                continue

            if hasattr(self.db, "get_searchable_mirror_hash"):
                if self.db.get_searchable_mirror_hash(str(path_obj)) == file_hash:
                    continue

            raw_text = self._extract_text(path_obj)
            if not raw_text:
                continue

            if hasattr(self.db, "save_searchable_mirror"):
                self.db.save_searchable_mirror(
                    file_path=str(path_obj),
                    raw_text=raw_text,
                    file_hash=file_hash,
                )

            self._log_index_event(
                {
                    "file_path": str(path_obj),
                    "file_hash": file_hash,
                    "size_bytes": path_obj.stat().st_size if path_obj.exists() else 0,
                    "indexed_at": time.strftime("%Y-%m-%d %H:%M:%S"),
                }
            )

            processed += 1
            if max_files is not None and processed >= max_files:
                break

            if sleep_sec > 0:
                time.sleep(sleep_sec)

        if processed > 0:
            try:
                train_root = Path(CONFIG.get("paths", {}).get("train_root", "D:/Train"))
                self.build_style_profile(train_root)
            except Exception:
                pass
        return processed


class IdleIndexController:
    def __init__(
        self,
        indexer: DocumentIndexer,
        idle_min_sec: int,
        idle_max_sec: int,
        max_files_per_cycle: int,
        sleep_between_files_sec: float,
        status_cb=None,
    ):
        self.indexer = indexer
        self.idle_min_sec = max(10, int(idle_min_sec))
        self.idle_max_sec = max(self.idle_min_sec, int(idle_max_sec))
        self.max_files_per_cycle = max(1, int(max_files_per_cycle))
        self.sleep_between_files_sec = float(sleep_between_files_sec)
        self.status_cb = status_cb
        self._running = False
        self._thread = None

    def start(self):
        if self._running:
            return
        self._running = True
        self._thread = threading.Thread(target=self._run, daemon=True)
        self._thread.start()

    def stop(self):
        self._running = False
        if self.indexer:
            self.indexer.stop()

    def _emit_status(self, message):
        if self.status_cb:
            try:
                self.status_cb(message)
            except Exception:
                pass

    @staticmethod
    def _get_idle_seconds():
        if os.name != "nt":
            return 0.0
        try:
            import ctypes

            class LASTINPUTINFO(ctypes.Structure):
                _fields_ = [("cbSize", ctypes.c_uint), ("dwTime", ctypes.c_uint)]

            info = LASTINPUTINFO()
            info.cbSize = ctypes.sizeof(info)
            if ctypes.windll.user32.GetLastInputInfo(ctypes.byref(info)) == 0:
                return 0.0
            tick_count = ctypes.windll.kernel32.GetTickCount()
            idle_ms = tick_count - info.dwTime
            return max(0.0, idle_ms / 1000.0)
        except Exception:
            return 0.0

    def _run(self):
        while self._running:
            idle_sec = self._get_idle_seconds()
            if idle_sec >= self.idle_min_sec:
                self._emit_status("Idle training: indexing documents...")
                processed = self.indexer.run_index_cycle(
                    max_files=self.max_files_per_cycle,
                    sleep_sec=self.sleep_between_files_sec,
                )
                if processed <= 0:
                    time.sleep(1.0)
                continue

            self._emit_status("Idle training: paused")
            time.sleep(2.0)


def build_default_indexer(db, status_cb=None):
    cfg = _default_training_config()
    train_root = cfg["train_root"]
    response_dir = cfg["response_dir"]
    include_exts = cfg["include_extensions"]
    skip_dirs = cfg["skip_dirs"]

    roots = [train_root, "C:/", "D:/"]
    indexer = DocumentIndexer(
        db=db,
        roots=roots,
        response_dir=response_dir,
        include_extensions=include_exts,
        skip_dirs=skip_dirs,
        status_cb=status_cb,
    )
    indexer.ensure_response_dir()

    controller = IdleIndexController(
        indexer=indexer,
        idle_min_sec=cfg["idle_min_sec"],
        idle_max_sec=cfg["idle_max_sec"],
        max_files_per_cycle=cfg["max_files_per_cycle"],
        sleep_between_files_sec=cfg["sleep_between_files_sec"],
        status_cb=status_cb,
    )
    return indexer, controller
