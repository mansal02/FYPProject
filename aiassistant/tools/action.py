import pyautogui
import re
import time
import os
import subprocess
import threading
import json
import platform
import shutil
import pywhatkit
import csv
import importlib
import random
import io
import contextlib
import webbrowser
from io import StringIO
from urllib.parse import quote_plus
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from aiassistant.infra.config.app_config import CONFIG
from aiassistant.tools.tools_os import run_tool_action

try:
    Document = importlib.import_module("docx").Document
    DOCX_AVAILABLE = True
except ImportError:
    Document = None
    DOCX_AVAILABLE = False

try:
    requests = importlib.import_module("requests")
    REQUESTS_AVAILABLE = True
except ImportError:
    requests = None
    REQUESTS_AVAILABLE = False

try:
    BeautifulSoup = importlib.import_module("bs4").BeautifulSoup
    BS4_AVAILABLE = True
except ImportError:
    BeautifulSoup = None
    BS4_AVAILABLE = False

try:
    pyperclip = importlib.import_module("pyperclip")
    CLIPBOARD_AVAILABLE = True
except ImportError:
    pyperclip = None
    CLIPBOARD_AVAILABLE = False

try:
    Presentation = importlib.import_module("pptx").Presentation
    POWERPOINT_AVAILABLE = True
except ImportError:
    Presentation = None
    POWERPOINT_AVAILABLE = False

SYSTEM_NAME = platform.system().lower()
IS_WINDOWS = SYSTEM_NAME.startswith("win")
TOOL_BRIDGE_VERBOSE = str(os.environ.get("MARIE_TOOL_BRIDGE_VERBOSE", "0")).strip().lower() in {
    "1",
    "true",
    "yes",
    "on",
}

GENERAL_TASK_COMMANDS = [
    "open <app>",
    "close <app>",
    "play <topic>",
    "write <text>",
    "note <text>",
    "type <text>",
    "take a note <text>",
    "volume up",
    "volume down",
    "mute",
    "unmute",
    "search web <query>",
    "open website <url>",
    "browse <url>",
    "research <query>",
    "web research <query>",
    "research web <query>",
]

CLIPBOARD_COMMANDS = [
    "copy selected text",
    "copy now",
    "paste clipboard",
    "paste now",
    "save clipboard to rad as <key>",
]

SYSTEM_COMMANDS = [
    "system check",
    "check system",
    "malware scan",
    "scan for malware",
    "run malware scan",
    "security quick scan",
    "scan apps",
    "update apps",
]

FILE_SYSTEM_COMMANDS = [
    "files roots",
    "files list <path>",
    "files deep search <query>",
    "files deep search <query> [in <root>]",
    "search file <hint>",
    "find file <hint>",
    "locate file <hint>",
    "files analyze <path>",
    "files create file <path> [content <text>]",
    "files create folder <path>",
    "files create dir <path>",
    "files create directory <path>",
    "files move <source> -> <destination>",
    "files copy <source> -> <destination>",
    "files delete <path>",
    "files remove <path>",
    "files open <path>",
    "open file <hint>",
]

SOFTWARE_COMMANDS = [
    "software open <app>",
    "software close <app>",
    "software running",
    "service open <gmail|outlook|whatsapp|telegram>",
    "service open <app_or_service>",
]

COMMUNICATION_COMMANDS = [
    "email draft to <email> subject <subject> body <body> [attach <file_hint>] [provider gmail|outlook|custom]",
    "email send to <email> subject <subject> body <body> [attach <file_hint>] [provider gmail|outlook|custom]",
    "telegram send to <chat_id> token <bot_token> message <text>",
    "telegram file <path_or_hint> to <chat_id> token <bot_token> [caption <text>]",
    "whatsapp send to <phone_number> message <text> (WhatsApp Web mode)",
]

ASSISTANT_TOOL_ACTIONS = [
    "list_system_roots",
    "list_directory",
    "deep_search (alias: deep_search_paths)",
    "analyze_path",
    "create_path",
    "move_path",
    "copy_path",
    "delete_path",
    "search_file",
    "semantic_search_file",
    "read_file",
    "open_file (alias: open_path)",
    "launch_application",
    "close_application",
    "list_running_apps (alias: list_running_applications)",
    "open_service",
    "search_mirror",
    "move_mouse",
    "click",
    "type_text",
    "press_key",
    "hotkey",
    "run_command",
    "toggle_dark_mode",
    "send_email",
    "draft_email_attachment",
    "send_telegram",
    "send_whatsapp",
    "online_query",
]

ASSISTANT_TOOL_EXAMPLES = [
    '<tool>{"action":"list_directory","path":"D:/pylearn/FYP/AiAssistant","max_results":120}</tool>',
    '<tool>{"action":"deep_search","query":"runsys.py","roots":["D:/pylearn/FYP/AiAssistant"],"include_content":false}</tool>',
    '<tool>{"action":"send_email","to":"name@example.com","subject":"Report","body":"Please find attached","provider":"outlook","attachments":["report.pdf"],"send_now":false}</tool>',
]

OFFICE_HELP_COMMANDS = [
    "office help",
    "excel help",
    "word help",
    "powerpoint help",
]

EXCEL_HELP_COMMANDS = [
    "create an excel workbook with <rows>x<cols> table with random data and then create the graph from it",
    "excel random table <rows>x<cols> with graph [in <file>]",
    "excel demo table graph [in <file>]",
    "excel create <file>",
    "excel create sheet <sheet_name> in <file>",
    "excel list sheets in <file>",
    "excel set <cell> to <value> in <file> [sheet <sheet_name>]",
    "excel get <cell> in <file> [sheet <sheet_name>]",
    "excel add row <comma-separated values> in <file> [sheet <sheet_name>]",
    "excel delete row <number> in <file> [sheet <sheet_name>]",
    "excel delete column <A..Z> in <file> [sheet <sheet_name>]",
    "excel sum <A1:B10> in <file> to <cell> [sheet <sheet_name>]",
    "excel formula <cell> = <formula> in <file> [sheet <sheet_name>]",
]

WORD_HELP_COMMANDS = [
    "word create <file>",
    "word add heading <text> [level 1-6] in <file>",
    "word add paragraph <text> in <file>",
    "word read <file>",
]

POWERPOINT_HELP_COMMANDS = [
    "powerpoint create <file>",
    "powerpoint add slide title <title> content <content> in <file>",
    "powerpoint launch <file>",
]

ASSISTANT_JSON_ACTIONS = [
    '{"action":"open","target":"chrome"}',
    '{"action":"close","target":"notepad"}',
    '{"action":"search_web","target":"latest ai news"}',
    '{"action":"open_website","target":"github.com"}',
    '{"action":"volume","target":"up"}',
    '{"action":"write_note","target":"meeting summary"}',
    '{"action":"play","target":"lofi coding music"}',
]


def _open_with_default_app(path):
    try:
        if IS_WINDOWS and hasattr(os, "startfile"):
            os.startfile(path)
        elif SYSTEM_NAME == "darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
        return True
    except Exception as e:
        print(f"[ACTION] Could not open file with default app: {e}")
        return False


try:
    # Import AppOpener functions (best on Windows).
    from AppOpener import open as open_app
    from AppOpener import close as close_app
    from AppOpener import give_appnames
    APPOPENER_AVAILABLE = True
except Exception:
    open_app = None
    close_app = None
    give_appnames = None
    APPOPENER_AVAILABLE = False

class ExcelCommandHandler:
    def __init__(self):
        self.default_extension = ".xlsx"

    def _normalize_path(self, file_name):
        clean_name = file_name.strip().strip('"').strip("'")
        if not clean_name.lower().endswith(self.default_extension):
            clean_name += self.default_extension

        is_windows_absolute = re.match(r"^[a-zA-Z]:\\", clean_name) is not None
        if not os.path.isabs(clean_name) and not is_windows_absolute:
            clean_name = os.path.abspath(clean_name)

        parent = os.path.dirname(clean_name)
        if parent:
            os.makedirs(parent, exist_ok=True)
        return clean_name

    def _load_or_create_workbook(self, path):
        if os.path.exists(path):
            return load_workbook(path)
        return Workbook()

    def _pick_sheet(self, workbook, sheet_name):
        if not sheet_name:
            return workbook.active
        target_sheet = sheet_name.strip().strip('"').strip("'")
        if target_sheet in workbook.sheetnames:
            return workbook[target_sheet]
        return workbook.create_sheet(target_sheet)

    def _to_value(self, raw_value):
        value = raw_value.strip()
        if (value.startswith('"') and value.endswith('"')) or (value.startswith("'") and value.endswith("'")):
            return value[1:-1]
        if value.lower() == "true":
            return True
        if value.lower() == "false":
            return False
        if re.fullmatch(r"-?\d+", value):
            return int(value)
        try:
            return float(value)
        except ValueError:
            pass
        return value

    def _ensure_formula_prefix(self, formula_text):
        cleaned_formula = formula_text.strip()
        return cleaned_formula if cleaned_formula.startswith("=") else "=" + cleaned_formula

    def _create_random_table_with_chart(self, file_name, rows=10, cols=10):
        file_path = self._normalize_path(file_name)
        workbook = self._load_or_create_workbook(file_path)
        sheet = self._pick_sheet(workbook, "RandomData")

        # Reset sheet contents when regenerating demo data.
        if sheet.max_row > 0:
            sheet.delete_rows(1, sheet.max_row)
        if sheet.max_column > 0:
            sheet.delete_cols(1, sheet.max_column)
        sheet._charts = []

        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                sheet.cell(row=row, column=col, value=random.randint(10, 99))

        chart = LineChart()
        chart.title = f"Random Data {rows}x{cols}"
        chart.style = 10
        chart.y_axis.title = "Value"
        chart.x_axis.title = "Row"

        data_ref = Reference(sheet, min_col=1, min_row=1, max_col=cols, max_row=rows)
        category_ref = Reference(sheet, min_col=1, min_row=1, max_row=rows)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(category_ref)

        chart_anchor = f"{get_column_letter(cols + 2)}2"
        sheet.add_chart(chart, chart_anchor)

        workbook.save(file_path)
        print(f"[ACTION][EXCEL] Random {rows}x{cols} table + chart created in {file_path}")

        try:
            if _open_with_default_app(file_path):
                print("[ACTION][EXCEL] Opened workbook in available spreadsheet software.")
            else:
                print("[ACTION][EXCEL] Workbook created, but opening the file failed.")
        except Exception as e:
            print(f"[ACTION][EXCEL] Workbook created but auto-open failed: {e}")

    def handle(self, text):
        random_nl_match = re.fullmatch(
            r"(?:create|make)\s+an?\s+excel\s+workbook\s+with\s+(\d+)x(\d+)\s+table\s+with\s+random\s+data\s+and\s+then\s+create\s+(?:the\s+)?graph\s+from\s+it(?:\s+in\s+available\s+software)?(?:\s+in\s+(.+?))?\s*$",
            text,
            flags=re.IGNORECASE,
        )
        if random_nl_match:
            rows, cols, file_name = random_nl_match.groups()
            rows = int(rows)
            cols = int(cols)
            if rows > 100 or cols > 50 or rows <= 0 or cols <= 0:
                print("[ACTION][EXCEL] Table size out of supported range. Use 1..100 rows and 1..50 columns.")
                return True
            target_file = file_name or "random_chart_demo.xlsx"
            self._create_random_table_with_chart(target_file, rows=rows, cols=cols)
            return True

        random_cmd_match = re.fullmatch(
            r"excel\s+random\s+table\s+(\d+)x(\d+)\s+with\s+graph(?:\s+in\s+(.+?))?\s*$",
            text,
            flags=re.IGNORECASE,
        )
        if random_cmd_match:
            rows, cols, file_name = random_cmd_match.groups()
            rows = int(rows)
            cols = int(cols)
            if rows > 100 or cols > 50 or rows <= 0 or cols <= 0:
                print("[ACTION][EXCEL] Table size out of supported range. Use 1..100 rows and 1..50 columns.")
                return True
            target_file = file_name or "random_chart_demo.xlsx"
            self._create_random_table_with_chart(target_file, rows=rows, cols=cols)
            return True

        if re.fullmatch(r"excel\s+demo\s+table\s+graph(?:\s+in\s+(.+?))?\s*$", text, flags=re.IGNORECASE):
            demo_file_match = re.fullmatch(r"excel\s+demo\s+table\s+graph(?:\s+in\s+(.+?))?\s*$", text, flags=re.IGNORECASE)
            target_file = demo_file_match.group(1) if demo_file_match else "random_chart_demo.xlsx"
            self._create_random_table_with_chart(target_file or "random_chart_demo.xlsx", rows=10, cols=10)
            return True

        open_match = re.fullmatch(r"excel\s+(?:create|open)\s+(.+?)\s*$", text, flags=re.IGNORECASE)
        if open_match:
            file_path = self._normalize_path(open_match.group(1))
            workbook = self._load_or_create_workbook(file_path)
            workbook.save(file_path)
            print(f"[ACTION][EXCEL] Ready: {file_path}")
            return True

        create_sheet_match = re.fullmatch(
            r"excel\s+create\s+sheet\s+(.+?)\s+in\s+(.+?)\s*$",
            text,
            flags=re.IGNORECASE,
        )
        if create_sheet_match:
            sheet_name, file_name = create_sheet_match.groups()
            file_path = self._normalize_path(file_name)
            workbook = self._load_or_create_workbook(file_path)
            target_sheet = sheet_name.strip().strip('"').strip("'")
            if target_sheet not in workbook.sheetnames:
                workbook.create_sheet(target_sheet)
                workbook.save(file_path)
                print(f"[ACTION][EXCEL] Sheet '{target_sheet}' created in {file_path}")
            else:
                print(f"[ACTION][EXCEL] Sheet '{target_sheet}' already exists in {file_path}")
            return True

        list_sheets_match = re.fullmatch(
            r"excel\s+list\s+sheets\s+in\s+(.+?)\s*$",
            text,
            flags=re.IGNORECASE,
        )
        if list_sheets_match:
            file_name = list_sheets_match.group(1)
            file_path = self._normalize_path(file_name)
            workbook = self._load_or_create_workbook(file_path)
            print(f"[ACTION][EXCEL] Sheets in {file_path}: {', '.join(workbook.sheetnames)}")
            return True

        set_match = re.fullmatch(
            r"excel\s+set\s+([a-zA-Z]+\d+)\s+to\s+(.+?)\s+in\s+(.+?)(?:\s+sheet\s+(.+))?",
            text,
            flags=re.IGNORECASE,
        )
        if set_match:
            cell, value_text, file_name, sheet_name = set_match.groups()
            file_path = self._normalize_path(file_name)
            workbook = self._load_or_create_workbook(file_path)
            sheet = self._pick_sheet(workbook, sheet_name)
            sheet[cell.upper()] = self._to_value(value_text)
            workbook.save(file_path)
            print(f"[ACTION][EXCEL] {cell.upper()} updated in {file_path}")
            return True

        get_match = re.fullmatch(
            r"excel\s+get\s+([a-zA-Z]+\d+)\s+in\s+(.+?)(?:\s+sheet\s+(.+))?",
            text,
            flags=re.IGNORECASE,
        )
        if get_match:
            cell, file_name, sheet_name = get_match.groups()
            file_path = self._normalize_path(file_name)
            workbook = self._load_or_create_workbook(file_path)
            sheet = self._pick_sheet(workbook, sheet_name)
            _ = sheet[cell.upper()].value
            print(f"[ACTION][EXCEL] {cell.upper()} retrieved from {file_path}. Content hidden by policy.")
            return True

        row_match = re.fullmatch(
            r"excel\s+add\s+row\s+(.+?)\s+in\s+(.+?)(?:\s+sheet\s+(.+))?",
            text,
            flags=re.IGNORECASE,
        )
        if row_match:
            row_values_text, file_name, sheet_name = row_match.groups()
            file_path = self._normalize_path(file_name)
            workbook = self._load_or_create_workbook(file_path)
            sheet = self._pick_sheet(workbook, sheet_name)
            try:
                parsed_row = next(csv.reader(StringIO(row_values_text)), [])
            except csv.Error:
                print("[ACTION][EXCEL] Row input is malformed. Use comma-separated values, e.g. apple,10,done.")
                return True
            if not parsed_row or all(not item.strip() for item in parsed_row):
                print("[ACTION][EXCEL] No valid row values provided. Use comma-separated values.")
                return True
            sheet.append([self._to_value(item.strip()) for item in parsed_row])
            workbook.save(file_path)
            print(f"[ACTION][EXCEL] Row added in {file_path}")
            return True

        delete_row_match = re.fullmatch(
            r"excel\s+delete\s+row\s+(\d+)\s+in\s+(.+?)(?:\s+sheet\s+(.+))?",
            text,
            flags=re.IGNORECASE,
        )
        if delete_row_match:
            row_number, file_name, sheet_name = delete_row_match.groups()
            row_number = int(row_number)
            if row_number <= 0:
                print("[ACTION][EXCEL] Row number must be 1 or greater.")
                return True
            file_path = self._normalize_path(file_name)
            workbook = self._load_or_create_workbook(file_path)
            sheet = self._pick_sheet(workbook, sheet_name)
            sheet.delete_rows(row_number, 1)
            workbook.save(file_path)
            print(f"[ACTION][EXCEL] Row {row_number} deleted in {file_path}")
            return True

        delete_col_match = re.fullmatch(
            r"excel\s+delete\s+column\s+([a-zA-Z]+)\s+in\s+(.+?)(?:\s+sheet\s+(.+))?",
            text,
            flags=re.IGNORECASE,
        )
        if delete_col_match:
            col_letters, file_name, sheet_name = delete_col_match.groups()
            file_path = self._normalize_path(file_name)
            workbook = self._load_or_create_workbook(file_path)
            sheet = self._pick_sheet(workbook, sheet_name)
            col_letters = col_letters.upper()
            col_index = 0
            for ch in col_letters:
                col_index = col_index * 26 + (ord(ch) - ord("A") + 1)
            sheet.delete_cols(col_index, 1)
            workbook.save(file_path)
            print(f"[ACTION][EXCEL] Column {col_letters} deleted in {file_path}")
            return True

        sum_match = re.fullmatch(
            r"excel\s+sum\s+([a-zA-Z]+\d+:[a-zA-Z]+\d+)\s+in\s+(.+?)\s+to\s+([a-zA-Z]+\d+)(?:\s+sheet\s+(.+))?",
            text,
            flags=re.IGNORECASE,
        )
        if sum_match:
            source_range, file_name, target_cell, sheet_name = sum_match.groups()
            file_path = self._normalize_path(file_name)
            workbook = self._load_or_create_workbook(file_path)
            sheet = self._pick_sheet(workbook, sheet_name)
            sheet[target_cell.upper()] = self._ensure_formula_prefix(f"SUM({source_range.upper()})")
            workbook.save(file_path)
            print(f"[ACTION][EXCEL] Sum {source_range.upper()} -> {target_cell.upper()} in {file_path}")
            return True

        formula_match = re.fullmatch(
            r"excel\s+formula\s+([a-zA-Z]+\d+)\s*=\s*(.+?)\s+in\s+(.+?)(?:\s+sheet\s+(.+))?",
            text,
            flags=re.IGNORECASE,
        )
        if formula_match:
            target_cell, formula_text, file_name, sheet_name = formula_match.groups()
            file_path = self._normalize_path(file_name)
            workbook = self._load_or_create_workbook(file_path)
            sheet = self._pick_sheet(workbook, sheet_name)
            sheet[target_cell.upper()] = self._ensure_formula_prefix(formula_text)
            workbook.save(file_path)
            print(f"[ACTION][EXCEL] Formula set in {target_cell.upper()} for {file_path}")
            return True

        if re.fullmatch(r"excel\s+help", text, flags=re.IGNORECASE):
            print("[ACTION][EXCEL] Commands:")
            for command in EXCEL_HELP_COMMANDS:
                print(f"- {command}")
            return True

        return False


class WordCommandHandler:
    def __init__(self):
        self.default_extension = ".docx"

    def _normalize_path(self, file_name):
        clean_name = file_name.strip().strip('"').strip("'")
        if not clean_name.lower().endswith(self.default_extension):
            clean_name += self.default_extension

        is_windows_absolute = re.match(r"^[a-zA-Z]:\\", clean_name) is not None
        if not os.path.isabs(clean_name) and not is_windows_absolute:
            clean_name = os.path.abspath(clean_name)

        parent = os.path.dirname(clean_name)
        if parent:
            os.makedirs(parent, exist_ok=True)
        return clean_name

    def _load_or_create_doc(self, path):
        if not DOCX_AVAILABLE:
            return None
        if os.path.exists(path):
            return Document(path)
        return Document()

    def handle(self, text):
        if not DOCX_AVAILABLE:
            if text.lower().startswith("word "):
                print("[ACTION][WORD] python-docx is not installed. Run: pip install python-docx")
                return True
            return False

        create_match = re.fullmatch(r"word\s+(?:create|open)\s+(.+?)\s*$", text, flags=re.IGNORECASE)
        if create_match:
            file_path = self._normalize_path(create_match.group(1))
            doc = self._load_or_create_doc(file_path)
            doc.save(file_path)
            print(f"[ACTION][WORD] Ready: {file_path}")
            return True

        paragraph_match = re.fullmatch(
            r"word\s+add\s+paragraph\s+(.+?)\s+in\s+(.+?)\s*$",
            text,
            flags=re.IGNORECASE,
        )
        if paragraph_match:
            paragraph_text, file_name = paragraph_match.groups()
            file_path = self._normalize_path(file_name)
            doc = self._load_or_create_doc(file_path)
            content = paragraph_text.strip().strip('"').strip("'")
            doc.add_paragraph(content)
            doc.save(file_path)
            print(f"[ACTION][WORD] Paragraph added in {file_path}")
            return True

        heading_match = re.fullmatch(
            r"word\s+add\s+heading\s+(.+?)(?:\s+level\s+([1-6]))?\s+in\s+(.+?)\s*$",
            text,
            flags=re.IGNORECASE,
        )
        if heading_match:
            heading_text, level_text, file_name = heading_match.groups()
            level = int(level_text) if level_text else 1
            file_path = self._normalize_path(file_name)
            doc = self._load_or_create_doc(file_path)
            content = heading_text.strip().strip('"').strip("'")
            doc.add_heading(content, level=level)
            doc.save(file_path)
            print(f"[ACTION][WORD] Heading (level {level}) added in {file_path}")
            return True

        read_match = re.fullmatch(r"word\s+read\s+(.+?)\s*$", text, flags=re.IGNORECASE)
        if read_match:
            file_path = self._normalize_path(read_match.group(1))
            if not os.path.exists(file_path):
                print(f"[ACTION][WORD] File not found: {file_path}")
                return True
            doc = self._load_or_create_doc(file_path)
            _ = doc
            print(f"[ACTION][WORD] Document read from {file_path}. Content hidden by policy.")
            return True

        if re.fullmatch(r"word\s+help", text, flags=re.IGNORECASE):
            print("[ACTION][WORD] Commands:")
            for command in WORD_HELP_COMMANDS:
                print(f"- {command}")
            return True

        return False


class PowerPointCommandHandler:
    def __init__(self):
        self.default_extension = ".pptx"

    def _normalize_path(self, file_name):
        clean_name = file_name.strip().strip('"').strip("'")
        if not clean_name.lower().endswith(self.default_extension):
            clean_name += self.default_extension

        is_windows_absolute = re.match(r"^[a-zA-Z]:\\", clean_name) is not None
        if not os.path.isabs(clean_name) and not is_windows_absolute:
            clean_name = os.path.abspath(clean_name)

        parent = os.path.dirname(clean_name)
        if parent:
            os.makedirs(parent, exist_ok=True)
        return clean_name

    def _load_or_create_presentation(self, path):
        if not POWERPOINT_AVAILABLE:
            return None
        if os.path.exists(path):
            return Presentation(path)
        return Presentation()

    def handle(self, text):
        if not POWERPOINT_AVAILABLE:
            if text.lower().startswith("powerpoint "):
                print("[ACTION][PPT] python-pptx is not installed. Run: pip install python-pptx")
                return True
            return False

        create_match = re.fullmatch(r"powerpoint\s+(?:create|open)\s+(.+?)\s*$", text, flags=re.IGNORECASE)
        if create_match:
            file_path = self._normalize_path(create_match.group(1))
            prs = self._load_or_create_presentation(file_path)
            prs.save(file_path)
            print(f"[ACTION][PPT] Ready: {file_path}")
            return True

        slide_match = re.fullmatch(
            r"powerpoint\s+add\s+slide\s+title\s+(.+?)\s+content\s+(.+?)\s+in\s+(.+?)\s*$",
            text,
            flags=re.IGNORECASE,
        )
        if slide_match:
            title_text, content_text, file_name = slide_match.groups()
            file_path = self._normalize_path(file_name)
            prs = self._load_or_create_presentation(file_path)
            layout = prs.slide_layouts[1] if len(prs.slide_layouts) > 1 else prs.slide_layouts[0]
            slide = prs.slides.add_slide(layout)

            title_box = getattr(slide.shapes, "title", None)
            if title_box:
                title_box.text = title_text.strip().strip('"').strip("'")

            if len(slide.placeholders) > 1:
                slide.placeholders[1].text = content_text.strip().strip('"').strip("'")

            prs.save(file_path)
            print(f"[ACTION][PPT] Slide added in {file_path}")
            return True

        launch_match = re.fullmatch(r"powerpoint\s+launch\s+(.+?)\s*$", text, flags=re.IGNORECASE)
        if launch_match:
            file_path = self._normalize_path(launch_match.group(1))
            if not os.path.exists(file_path):
                print(f"[ACTION][PPT] File not found: {file_path}")
                return True
            try:
                if _open_with_default_app(file_path):
                    print(f"[ACTION][PPT] Opened {file_path}")
                else:
                    print(f"[ACTION][PPT] Failed to open {file_path}")
            except Exception as e:
                print(f"[ACTION][PPT] Failed to open: {e}")
            return True

        if re.fullmatch(r"powerpoint\s+help", text, flags=re.IGNORECASE):
            print("[ACTION][PPT] Commands:")
            for command in POWERPOINT_HELP_COMMANDS:
                print(f"- {command}")
            return True

        return False

class ActionHandler:
    @staticmethod
    def get_supported_command_sections():
        return {
            "General task commands": list(GENERAL_TASK_COMMANDS),
            "File system commands": list(FILE_SYSTEM_COMMANDS),
            "Software control commands": list(SOFTWARE_COMMANDS),
            "Communication commands": list(COMMUNICATION_COMMANDS),
            "Clipboard commands": list(CLIPBOARD_COMMANDS),
            "System commands": list(SYSTEM_COMMANDS),
            "Office quick help": list(OFFICE_HELP_COMMANDS),
            "Excel commands": list(EXCEL_HELP_COMMANDS),
            "Word commands": list(WORD_HELP_COMMANDS),
            "PowerPoint commands": list(POWERPOINT_HELP_COMMANDS),
            "Assistant JSON command format": list(ASSISTANT_JSON_ACTIONS),
            "Assistant tool actions (action field)": list(ASSISTANT_TOOL_ACTIONS),
            "Assistant tool call examples": list(ASSISTANT_TOOL_EXAMPLES),
        }

    def __init__(self, db=None, context_provider=None):
        # 1. CUSTOM APPS / GAMES
        # Add games or portable apps here that the scanner misses.
        # Use double backslashes \\ for paths.
        self.custom_apps = {
   
            "steam": r"C:\Program Files (x86)\Steam\steam.exe",
            "obs": r"C:\Program Files\obs-studio\bin\64bit\obs64.exe"
        }
        self.db = db
        self.context_provider = context_provider
        self.excel = ExcelCommandHandler()
        self.word = WordCommandHandler()
        self.powerpoint = PowerPointCommandHandler()
        action_cfg = CONFIG.get("actions", {})
        self.safe_mode = bool(action_cfg.get("safe_mode", True))
        self.allow_legacy_text_commands = bool(action_cfg.get("allow_legacy_text_commands", True))

    def execute_and_collect(self, text):
        if not text:
            return ""

        buffer = io.StringIO()
        with contextlib.redirect_stdout(buffer):
            try:
                self.execute(text)
            except Exception as e:
                print(f"[ACTION ERROR] {e}")
        return buffer.getvalue().strip()

    def _looks_like_action_command(self, text):
        if not text:
            return False
        lowered = text.strip().lower()
        prefixes = (
            "files ",
            "software ",
            "service ",
            "email ",
            "telegram ",
            "whatsapp ",
            "search file ",
            "find file ",
            "locate file ",
            "open file ",
            "open folder ",
            "open document ",
            "open ",
            "close ",
            "play ",
            "write ",
            "note ",
            "type ",
            "take a note ",
            "search web ",
            "open website ",
            "browse ",
            "excel ",
            "word ",
            "powerpoint ",
            "research ",
            "web research ",
            "system check",
            "check system",
            "scan apps",
            "update apps",
            "copy selected text",
            "copy now",
            "paste clipboard",
            "paste now",
            "save clipboard to rad as ",
            "malware scan",
            "scan for malware",
            "run malware scan",
            "security quick scan",
        )
        if lowered.startswith(prefixes):
            return True
        if re.search(r"\b(?:open|close)\b\s+[a-z0-9]", lowered):
            return True
        if re.search(r"\b(?:open|search|find|locate|look\s+for)\b.*\b(?:file|files|folder|folders|document|documents|path)\b", lowered):
            return True
        return any(token in lowered for token in ("volume up", "volume down", "mute", "unmute"))

    @staticmethod
    def _looks_like_file_hint(raw_text):
        clean = str(raw_text or "").strip().lower()
        if not clean:
            return False

        if re.search(r"^[a-z]:\\", clean):
            return True
        if "\\" in clean or "/" in clean:
            return True
        if re.search(r"\.[a-z0-9]{1,6}\b", clean):
            return True

        file_tokens = (
            " file",
            "files ",
            "folder",
            "document",
            "directory",
            "path",
            "report",
            "notes",
            "draft",
            "summary",
            "invoice",
            "project",
        )
        return any(token in clean for token in file_tokens)

    def _print_system_check(self):
        checks = {
            "Python": True,
            "Reasoning Server Module": importlib.util.find_spec("aiassistant.backend.server_reasoning") is not None,
            "Voice Server Module": importlib.util.find_spec("aiassistant.backend.server_voice") is not None,
            "Piper Folder": os.path.isdir("piper"),
            "Models Folder": os.path.isdir("models"),
            "RVC Models Folder": os.path.isdir("rvc_models"),
            "Internet Library (requests)": REQUESTS_AVAILABLE,
            "Clipboard Library (pyperclip)": CLIPBOARD_AVAILABLE,
        }

        print("[ACTION][SYSTEM] Quick Check:")
        for name, is_ok in checks.items():
            state = "OK" if is_ok else "MISSING"
            print(f"- {name}: {state}")

    def _run_quick_malware_scan(self):
        print("[ACTION][SECURITY] Starting Windows Defender quick scan...")
        try:
            subprocess.run(
                [
                    "powershell",
                    "-NoProfile",
                    "-Command",
                    "Start-MpScan -ScanType QuickScan",
                ],
                check=False,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
            print("[ACTION][SECURITY] Quick scan request sent to Windows Defender.")
        except Exception as e:
            print(f"[ACTION][SECURITY] Could not start Defender scan: {e}")

    def _open_website(self, target):
        url = target.strip().strip('"').strip("'")
        if not url.startswith(("http://", "https://")):
            url = "https://" + url
        webbrowser.open(url)
        print(f"[ACTION][WEB] Opened {url}")

    def _extract_key_points(self, text, max_points=5):
        if not text:
            return []

        chunks = re.split(r"(?<=[.!?])\s+", text)
        points = []
        for chunk in chunks:
            cleaned = re.sub(r"\s+", " ", chunk).strip(" -\n\t")
            if len(cleaned) < 30:
                continue
            if cleaned in points:
                continue
            points.append(cleaned)
            if len(points) >= max_points:
                break
        return points

    def _store_web_points_to_rad(self, query, points, source_url=""):
        if not self.db or not hasattr(self.db, "add_rad_data_if_new"):
            return 0

        slug = re.sub(r"[^a-z0-9]+", "_", query.lower()).strip("_")[:32] or "query"
        stored_count = 0

        for idx, point in enumerate(points, start=1):
            key = f"web_{slug}_{idx}"
            if self.db.add_rad_data_if_new("web_fact", key, point, 0.78):
                stored_count += 1

        if source_url:
            if self.db.add_rad_data_if_new("web_source", f"source_{slug}", source_url, 0.95):
                stored_count += 1

        return stored_count

    def _fetch_web_text(self, url):
        if not REQUESTS_AVAILABLE:
            return ""

        response = requests.get(
            url,
            headers={"User-Agent": "Mozilla/5.0 (compatible; MARIE/1.0)"},
            timeout=12,
        )
        response.raise_for_status()

        html_text = response.text
        if BS4_AVAILABLE:
            soup = BeautifulSoup(html_text, "html.parser")
            paragraphs = [p.get_text(" ", strip=True) for p in soup.find_all("p")]
            joined = " ".join(paragraphs)
            return re.sub(r"\s+", " ", joined).strip()

        no_scripts = re.sub(r"<script[\s\S]*?</script>", " ", html_text, flags=re.IGNORECASE)
        no_styles = re.sub(r"<style[\s\S]*?</style>", " ", no_scripts, flags=re.IGNORECASE)
        plain = re.sub(r"<[^>]+>", " ", no_styles)
        return re.sub(r"\s+", " ", plain).strip()

    def _research_and_store_web_data(self, query):
        if not REQUESTS_AVAILABLE:
            print("[ACTION][WEB] requests is not installed. Run: pip install requests")
            return

        query = query.strip()
        if not query:
            print("[ACTION][WEB] Missing query. Example: web research latest AI trends")
            return

        source_url = ""
        combined_text = ""

        try:
            if query.startswith(("http://", "https://")):
                source_url = query
                combined_text = self._fetch_web_text(source_url)
            else:
                ddg_url = "https://api.duckduckgo.com/"
                response = requests.get(
                    ddg_url,
                    params={
                        "q": query,
                        "format": "json",
                        "no_html": 1,
                        "skip_disambig": 1,
                    },
                    timeout=10,
                )
                response.raise_for_status()
                data = response.json()

                abstract = data.get("AbstractText", "")
                source_url = data.get("AbstractURL", "")
                related_lines = []
                for item in data.get("RelatedTopics", [])[:10]:
                    if isinstance(item, dict) and item.get("Text"):
                        related_lines.append(item["Text"])
                    elif isinstance(item, dict) and item.get("Topics"):
                        for sub in item.get("Topics", [])[:5]:
                            if isinstance(sub, dict) and sub.get("Text"):
                                related_lines.append(sub["Text"])

                combined_text = " ".join([abstract] + related_lines)

                if source_url:
                    fetched_page_text = self._fetch_web_text(source_url)
                    if fetched_page_text:
                        combined_text = f"{combined_text} {fetched_page_text}"

            points = self._extract_key_points(combined_text, max_points=6)
            if not points:
                print("[ACTION][WEB] Could not extract useful points from the web result.")
                return

            print(f"[ACTION][WEB] Key points for '{query}':")
            for idx, point in enumerate(points, start=1):
                print(f"{idx}. {point}")

            stored_count = self._store_web_points_to_rad(query, points, source_url)
            if stored_count > 0:
                print(f"[ACTION][RAD] Stored {stored_count} web-derived memory items.")

            if CLIPBOARD_AVAILABLE:
                pyperclip.copy("\n".join(points))
                print("[ACTION][WEB] Copied key points to clipboard.")
        except Exception as e:
            print(f"[ACTION][WEB] Research failed: {e}")

    def _open_text_editor(self):
        try:
            if IS_WINDOWS:
                os.system("start notepad")
                return
            if SYSTEM_NAME == "darwin":
                subprocess.Popen(["open", "-a", "TextEdit"])
                return

            for editor in ("gedit", "kate", "mousepad", "xed"):
                if shutil.which(editor):
                    subprocess.Popen([editor])
                    return
        except Exception as e:
            print(f"[ACTION] Could not open text editor: {e}")

    def _open_app_name(self, app_name):
        print(f"[ACTION] Opening: '{app_name}'")

        for key, path in self.custom_apps.items():
            if key in app_name:
                print(f"[ACTION] Found custom path for {key}")
                try:
                    if _open_with_default_app(path):
                        return True
                except Exception as e:
                    print(f"[ERROR] Custom path failed: {e}")

        if APPOPENER_AVAILABLE and IS_WINDOWS:
            try:
                open_app(app_name, match_closest=True, output=False, throw_error=True)
                return True
            except Exception:
                pass

        try:
            if IS_WINDOWS:
                os.system(f"start {app_name}")
            elif SYSTEM_NAME == "darwin":
                subprocess.Popen(["open", "-a", app_name])
            else:
                subprocess.Popen([app_name])
            return True
        except Exception:
            print(f"[ERROR] Could not open '{app_name}'")
            return False

    def _close_app_name(self, app_name):
        if APPOPENER_AVAILABLE and IS_WINDOWS:
            try:
                close_app(app_name, match_closest=True, output=False, throw_error=True)
                return True
            except Exception:
                pass

        try:
            if IS_WINDOWS:
                subprocess.run(["taskkill", "/f", "/im", f"{app_name}.exe"], check=False)
            else:
                subprocess.run(["pkill", "-f", app_name], check=False)
            return True
        except Exception:
            print(f"[ERROR] Could not close '{app_name}'")
            return False

    def _extract_json_action(self, text):
        if not text:
            return None

        candidates = []
        stripped = text.strip()
        if stripped.startswith("{") and stripped.endswith("}"):
            candidates.append(stripped)

        fenced = re.findall(r"```json\s*(\{[\s\S]*?\})\s*```", text, flags=re.IGNORECASE)
        candidates.extend(fenced)

        loose = re.findall(r"(\{\s*\"action\"[\s\S]*?\})", text)
        candidates.extend(loose)

        for candidate in candidates:
            try:
                obj = json.loads(candidate)
            except Exception:
                continue

            if not isinstance(obj, dict):
                continue

            action = str(obj.get("action", "")).strip().lower()
            target = str(obj.get("target", "")).strip()
            value = obj.get("value", "")
            if action in {"open", "close", "search_web", "open_website", "volume", "write_note", "play"}:
                return {"action": action, "target": target, "value": value}

        return None

    def _print_tool_bridge_result(self, result):
        if not isinstance(result, dict):
            print(f"[ACTION][TOOLS] {result}")
            return False

        ok = bool(result.get("success", False))
        prefix = "[ACTION][TOOLS]" if ok else "[ACTION][TOOLS][ERROR]"
        message = str(result.get("message", "")).strip()
        low_message = message.lower()
        hide_zero_match = "found 0 lexical file match" in low_message
        if message and not hide_zero_match:
            print(f"{prefix} {message}")

        payload = result.get("data")
        payload_has_value = payload not in (None, [], {}, "")
        if TOOL_BRIDGE_VERBOSE and payload_has_value:
            try:
                text = json.dumps(payload, ensure_ascii=True)
            except Exception:
                text = str(payload)
            if len(text) > 1800:
                text = text[:1800] + "..."
            print(f"{prefix} Data: {text}")

        error = str(result.get("error", "")).strip()
        if error and not ok:
            print(f"{prefix} {error}")
        return ok

    def _run_tool_bridge_action(self, action):
        try:
            result = run_tool_action(action)
            return self._print_tool_bridge_result(result)
        except Exception as e:
            print(f"[ACTION][TOOLS][ERROR] {e}")
            return False

    def _handle_extended_tool_commands(self, raw_text):
        if not raw_text:
            return False

        lowered = raw_text.strip().lower()

        natural_search_match = re.search(
            r"\b(?:search|find|locate|look\s+for)\b.*?\b(?:file|files|folder|folders|document|documents|path)\b\s+(.+)$",
            raw_text,
            flags=re.IGNORECASE,
        )
        if natural_search_match and "website" not in lowered and "search web" not in lowered:
            hint = natural_search_match.group(1).strip().strip('"').strip("'")
            if hint:
                self._run_tool_bridge_action(
                    {
                        "action": "search_file",
                        "hint": hint,
                        "max_results": 20,
                        "include_content": True,
                    }
                )
                return True

        natural_open_match = re.search(
            r"\bopen\b.*?\b(?:file|folder|document|path)\b\s+(.+)$",
            raw_text,
            flags=re.IGNORECASE,
        )
        if natural_open_match and "website" not in lowered:
            target_hint = natural_open_match.group(1).strip().strip('"').strip("'")
            if target_hint:
                self._run_tool_bridge_action(
                    {
                        "action": "open_file",
                        "path": target_hint,
                        "resolve_by_hint": True,
                        "include_content": True,
                    }
                )
                return True

        # Conversational app open/launch requests.
        natural_open_app_match = re.search(r"\b(?:open|launch|start|run)\b\s+(.+)$", raw_text, flags=re.IGNORECASE)
        if natural_open_app_match:
            if not re.search(r"\b(file|files|folder|folders|document|documents|path)\b", lowered):
                app_target = natural_open_app_match.group(1).strip().strip('"').strip("'")
                app_target = re.sub(
                    r"\b(?:please|now|for\s+me|thanks|thank\s+you)\b",
                    " ",
                    app_target,
                    flags=re.IGNORECASE,
                )
                app_target = re.sub(
                    r"\b(?:app|application|software|program)\b",
                    " ",
                    app_target,
                    flags=re.IGNORECASE,
                )
                app_target = re.sub(r"^(?:the|a|an)\s+", "", app_target, flags=re.IGNORECASE)
                app_target = re.sub(r"\s+", " ", app_target).strip()
                if app_target and "website" not in app_target.lower():
                    if re.search(r"(?:https?://|www\.|\.[a-z]{2,6}(?:/|$))", app_target, flags=re.IGNORECASE):
                        self._open_website(app_target)
                    else:
                        self._run_tool_bridge_action(
                            {
                                "action": "launch_application",
                                "app": app_target,
                            }
                        )
                    return True

        natural_close_app_match = re.search(
            r"\b(?:close|quit|exit|kill|terminate)\b\s+(.+)$",
            raw_text,
            flags=re.IGNORECASE,
        )
        if natural_close_app_match:
            if not re.search(r"\b(file|files|folder|folders|document|documents|path)\b", lowered):
                close_target = natural_close_app_match.group(1).strip().strip('"').strip("'")
                close_target = re.sub(
                    r"\b(?:please|now|for\s+me|thanks|thank\s+you)\b",
                    " ",
                    close_target,
                    flags=re.IGNORECASE,
                )
                close_target = re.sub(
                    r"\b(?:app|application|software|program)\b",
                    " ",
                    close_target,
                    flags=re.IGNORECASE,
                )
                close_target = re.sub(r"^(?:the|a|an)\s+", "", close_target, flags=re.IGNORECASE)
                close_target = re.sub(r"\s+", " ", close_target).strip()
                if close_target:
                    self._run_tool_bridge_action(
                        {
                            "action": "close_application",
                            "app": close_target,
                        }
                    )
                    return True

        # File system commands.
        if re.fullmatch(r"files\s+roots\s*", raw_text, flags=re.IGNORECASE):
            self._run_tool_bridge_action({"action": "list_system_roots"})
            return True

        match = re.fullmatch(r"files\s+list\s+(.+?)\s*$", raw_text, flags=re.IGNORECASE)
        if match:
            self._run_tool_bridge_action(
                {
                    "action": "list_directory",
                    "path": match.group(1).strip(),
                    "max_results": 120,
                }
            )
            return True

        match = re.fullmatch(r"files\s+deep\s+search\s+(.+?)(?:\s+in\s+(.+))?\s*$", raw_text, flags=re.IGNORECASE)
        if match:
            query, root = match.groups()
            action = {
                "action": "deep_search",
                "query": query.strip(),
                "max_results": 40,
                "include_content": True,
            }
            if root:
                action["roots"] = [root.strip()]
            else:
                action["include_all_drives"] = True
            self._run_tool_bridge_action(action)
            return True

        match = re.fullmatch(r"files\s+analyze\s+(.+?)\s*$", raw_text, flags=re.IGNORECASE)
        if match:
            self._run_tool_bridge_action(
                {
                    "action": "analyze_path",
                    "path": match.group(1).strip(),
                }
            )
            return True

        match = re.fullmatch(
            r"files\s+create\s+file\s+(.+?)(?:\s+content\s+(.+))?\s*$",
            raw_text,
            flags=re.IGNORECASE,
        )
        if match:
            path, content = match.groups()
            self._run_tool_bridge_action(
                {
                    "action": "create_path",
                    "path": path.strip(),
                    "kind": "file",
                    "content": (content or "").strip(),
                    "overwrite": False,
                }
            )
            return True

        match = re.fullmatch(r"files\s+create\s+(?:folder|dir|directory)\s+(.+?)\s*$", raw_text, flags=re.IGNORECASE)
        if match:
            self._run_tool_bridge_action(
                {
                    "action": "create_path",
                    "path": match.group(1).strip(),
                    "kind": "directory",
                }
            )
            return True

        match = re.fullmatch(r"files\s+move\s+(.+?)\s*->\s*(.+?)\s*$", raw_text, flags=re.IGNORECASE)
        if match:
            src, dst = match.groups()
            self._run_tool_bridge_action(
                {
                    "action": "move_path",
                    "src": src.strip(),
                    "dst": dst.strip(),
                }
            )
            return True

        match = re.fullmatch(r"files\s+copy\s+(.+?)\s*->\s*(.+?)\s*$", raw_text, flags=re.IGNORECASE)
        if match:
            src, dst = match.groups()
            self._run_tool_bridge_action(
                {
                    "action": "copy_path",
                    "src": src.strip(),
                    "dst": dst.strip(),
                }
            )
            return True

        match = re.fullmatch(r"files\s+(?:delete|remove)\s+(.+?)\s*$", raw_text, flags=re.IGNORECASE)
        if match:
            self._run_tool_bridge_action(
                {
                    "action": "delete_path",
                    "path": match.group(1).strip(),
                    "recursive": True,
                    "use_trash": True,
                }
            )
            return True

        match = re.fullmatch(r"files\s+open\s+(.+?)\s*$", raw_text, flags=re.IGNORECASE)
        if match:
            self._run_tool_bridge_action(
                {
                    "action": "open_file",
                    "path": match.group(1).strip(),
                    "resolve_by_hint": True,
                    "include_content": True,
                }
            )
            return True

        # Software commands.
        match = re.fullmatch(r"software\s+open\s+(.+?)\s*$", raw_text, flags=re.IGNORECASE)
        if match:
            self._run_tool_bridge_action(
                {
                    "action": "launch_application",
                    "app": match.group(1).strip(),
                }
            )
            return True

        match = re.fullmatch(r"software\s+close\s+(.+?)\s*$", raw_text, flags=re.IGNORECASE)
        if match:
            self._run_tool_bridge_action(
                {
                    "action": "close_application",
                    "app": match.group(1).strip(),
                }
            )
            return True

        if re.fullmatch(r"software\s+running\s*", raw_text, flags=re.IGNORECASE):
            self._run_tool_bridge_action({"action": "list_running_apps", "max_results": 120})
            return True

        match = re.fullmatch(r"service\s+open\s+(.+?)\s*$", raw_text, flags=re.IGNORECASE)
        if match:
            self._run_tool_bridge_action(
                {
                    "action": "open_service",
                    "service": match.group(1).strip(),
                }
            )
            return True

        # Email and messaging commands.
        match = re.fullmatch(
            r"email\s+(draft|send)\s+to\s+(.+?)\s+subject\s+(.+?)\s+body\s+(.+?)(?:\s+attach\s+(.+?))?(?:\s+provider\s+(gmail|outlook|custom))?\s*$",
            raw_text,
            flags=re.IGNORECASE,
        )
        if match:
            mode, to_email, subject, body, attachment, provider = match.groups()
            payload = {
                "action": "send_email",
                "to": to_email.strip(),
                "subject": subject.strip(),
                "body": body.strip(),
                "provider": (provider or ("outlook" if IS_WINDOWS else "gmail")).strip(),
                "send_now": mode.strip().lower() == "send",
            }
            if attachment:
                payload["attachments"] = [attachment.strip()]
            self._run_tool_bridge_action(payload)
            return True

        match = re.fullmatch(
            r"telegram\s+send\s+to\s+(.+?)\s+token\s+(.+?)\s+message\s+(.+)\s*$",
            raw_text,
            flags=re.IGNORECASE,
        )
        if match:
            chat_id, token, message = match.groups()
            self._run_tool_bridge_action(
                {
                    "action": "send_telegram",
                    "chat_id": chat_id.strip(),
                    "bot_token": token.strip(),
                    "message": message.strip(),
                }
            )
            return True

        match = re.fullmatch(
            r"telegram\s+file\s+(.+?)\s+to\s+(.+?)\s+token\s+(.+?)(?:\s+caption\s+(.+))?\s*$",
            raw_text,
            flags=re.IGNORECASE,
        )
        if match:
            file_hint, chat_id, token, caption = match.groups()
            self._run_tool_bridge_action(
                {
                    "action": "send_telegram",
                    "chat_id": chat_id.strip(),
                    "bot_token": token.strip(),
                    "file_hint": file_hint.strip(),
                    "message": (caption or "").strip(),
                }
            )
            return True

        match = re.fullmatch(r"whatsapp\s+send\s+to\s+(.+?)\s+message\s+(.+)\s*$", raw_text, flags=re.IGNORECASE)
        if match:
            to_number, message = match.groups()
            self._run_tool_bridge_action(
                {
                    "action": "send_whatsapp",
                    "to": to_number.strip(),
                    "message": message.strip(),
                }
            )
            return True

        return False

    def _execute_json_action(self, action_obj):
        action = action_obj.get("action", "")
        target = (action_obj.get("target") or "").strip()
        value = action_obj.get("value", "")

        if action == "open":
            return self._open_app_name(re.sub(r"[^\w\s]", "", target.lower()))

        if action == "close":
            return self._close_app_name(re.sub(r"[^\w\s]", "", target.lower()))

        if action == "search_web":
            if not target:
                return False
            search_url = f"https://duckduckgo.com/?q={quote_plus(target)}"
            webbrowser.open(search_url)
            print(f"[ACTION][WEB] Searching web for: {target}")
            return True

        if action == "open_website":
            if not target:
                return False
            self._open_website(target)
            return True

        if action == "volume":
            value_text = f"{target} {value}".lower()
            if "up" in value_text:
                pyautogui.press("volumeup")
                return True
            if "down" in value_text:
                pyautogui.press("volumedown")
                return True
            if "mute" in value_text or "unmute" in value_text:
                pyautogui.press("volumemute")
                return True
            return False

        if action == "write_note":
            content = target or str(value)
            if not content:
                return False
            self._open_text_editor()
            time.sleep(1.0)
            pyautogui.write(content, interval=0.05)
            print(f"[ACTION] Writing to editor: {content}")
            return True

        if action == "play":
            if not target:
                return False
            pywhatkit.playonyt(target)
            print(f"[ACTION] Playing on YouTube: {target}")
            return True

        return False

    def execute_from_assistant(self, text):
        action_obj = self._extract_json_action(text)
        if not action_obj:
            print("[ACTION][SAFE] Assistant output has no valid JSON action. Ignored.")
            return False
        return self._execute_json_action(action_obj)

    def execute(self, text):
        if not text: return
        raw_text = text.strip()
        text = raw_text.lower()

        structured_action = self._extract_json_action(raw_text)
        if structured_action:
            self._execute_json_action(structured_action)
            return

        if self._handle_extended_tool_commands(raw_text):
            return

        # =========================================================
        # EXCEL COMMANDS
        # =========================================================
        if self.excel.handle(raw_text):
            return

        # =========================================================
        # WORD / POWERPOINT COMMANDS
        # =========================================================
        if text.startswith("word ") and self.word.handle(raw_text):
            return

        if text.startswith("powerpoint ") and self.powerpoint.handle(raw_text):
            return

        if text == "office help":
            self.excel.handle("excel help")
            self.word.handle("word help")
            self.powerpoint.handle("powerpoint help")
            return

        # =========================================================
        # WEB + STUDY + CLIPBOARD
        # =========================================================
        web_research_match = re.fullmatch(r"(?:web\s+research|research\s+web|research)\s+(.+)", raw_text, flags=re.IGNORECASE)
        if web_research_match:
            query = web_research_match.group(1)
            self._research_and_store_web_data(query)
            return

        open_web_match = re.fullmatch(r"(?:open\s+website|browse)\s+(.+)", raw_text, flags=re.IGNORECASE)
        if open_web_match:
            self._open_website(open_web_match.group(1))
            return

        search_web_match = re.fullmatch(r"search\s+web\s+(.+)", raw_text, flags=re.IGNORECASE)
        if search_web_match:
            query = search_web_match.group(1).strip()
            search_url = f"https://duckduckgo.com/?q={quote_plus(query)}"
            webbrowser.open(search_url)
            print(f"[ACTION][WEB] Searching web for: {query}")
            return

        if re.fullmatch(r"(?:copy\s+selected\s+text|copy\s+now)", text):
            pyautogui.hotkey("ctrl", "c")
            time.sleep(0.2)
            if CLIPBOARD_AVAILABLE:
                snippet = pyperclip.paste().strip()
                preview = snippet[:120] + ("..." if len(snippet) > 120 else "")
                print(f"[ACTION][CLIPBOARD] Copied: {preview}")
            else:
                print("[ACTION][CLIPBOARD] Copied selection (clipboard preview unavailable).")
            return

        if re.fullmatch(r"(?:paste\s+clipboard|paste\s+now)", text):
            pyautogui.hotkey("ctrl", "v")
            print("[ACTION][CLIPBOARD] Pasted clipboard into active software.")
            return

        save_clip_match = re.fullmatch(r"save\s+clipboard\s+to\s+rad\s+as\s+(.+)", raw_text, flags=re.IGNORECASE)
        if save_clip_match:
            if not CLIPBOARD_AVAILABLE:
                print("[ACTION][RAD] pyperclip not installed. Run: pip install pyperclip")
                return
            if not self.db or not hasattr(self.db, "add_rad_data_if_new"):
                print("[ACTION][RAD] Database connection unavailable for clipboard save.")
                return

            key_name = save_clip_match.group(1).strip().strip('"').strip("'")
            clip_text = pyperclip.paste().strip()
            if not clip_text:
                print("[ACTION][RAD] Clipboard is empty.")
                return

            stored = self.db.add_rad_data_if_new("clipboard_note", key_name, clip_text, 0.84)
            if stored:
                print(f"[ACTION][RAD] Clipboard saved under key '{key_name}'.")
            else:
                print(f"[ACTION][RAD] Duplicate clipboard note ignored for key '{key_name}'.")
            return

        # =========================================================
        # SYSTEM CHECKING / OPTIONAL SECURITY
        # =========================================================
        if re.fullmatch(r"(?:system check|check system)", text):
            self._print_system_check()
            return

        if re.fullmatch(r"(?:malware scan|scan for malware|run malware scan|security quick scan)", text):
            threading.Thread(target=self._run_quick_malware_scan, daemon=True).start()
            return

        # =========================================================
        # 0. SPECIAL COMMAND: UPDATE APP LIST
        # =========================================================
        if "scan apps" in text or "update apps" in text:
            print("[ACTION] Scanning for new apps...")
            if APPOPENER_AVAILABLE and give_appnames:
                # Run this in a thread so it doesn't freeze MARIE
                threading.Thread(target=give_appnames, daemon=True).start()
            else:
                print("[ACTION] App scanning is only available when AppOpener is installed.")
            return

        # =========================================================
        # 1. YOUTUBE (Play Video)
        # =========================================================
        if text.startswith("play "):
            video_topic = text.replace("play ", "").replace("please", "").strip()
            print(f"[ACTION] Playing on YouTube: {video_topic}")
            try:
                pywhatkit.playonyt(video_topic)
            except Exception as e:
                print(f"[ERROR] YouTube failed: {e}")
            return

        # =========================================================
        # 2. NOTEPAD (Write text)
        # =========================================================
        write_triggers = ["write ", "note ", "type ", "take a note "]
        triggered_word = next((w for w in write_triggers if text.startswith(w)), None)

        if triggered_word:
            content = text.replace(triggered_word, "").strip()
            print(f"[ACTION] Writing to editor: {content}")
            self._open_text_editor()
            time.sleep(1.0)
            pyautogui.write(content, interval=0.05)
            return

        # =========================================================
        # 3. SYSTEM CONTROLS
        # =========================================================
        if "volume up" in text:
            pyautogui.press('volumeup')
            return
        elif "volume down" in text:
            pyautogui.press('volumedown')
            return
        elif "mute" in text or "unmute" in text:
            pyautogui.press('volumemute')
            return

        # =========================================================
        # 4. OPEN APPS (Custom + General)
        # =========================================================
        if text.startswith("open "):
            raw_name = text.replace("open ", "").strip()
            app_name = raw_name.replace("please", "").replace("now", "").strip()
            app_name = re.sub(r'[^\w\s]', '', app_name)

            if self._looks_like_file_hint(raw_name):
                self._run_tool_bridge_action(
                    {
                        "action": "open_file",
                        "path": raw_name,
                        "resolve_by_hint": True,
                        "include_content": True,
                    }
                )
                return

            opened = self._open_app_name(app_name)
            if not opened and len(app_name.split()) >= 3:
                self._run_tool_bridge_action(
                    {
                        "action": "open_file",
                        "path": raw_name,
                        "resolve_by_hint": True,
                        "include_content": True,
                    }
                )
            return

        # =========================================================
        # 5. CLOSE APPS
        # =========================================================
        if text.startswith("close "):
            app_name = text.replace("close ", "").replace("please", "").strip()
            app_name = re.sub(r'[^\w\s]', '', app_name)

            self._close_app_name(app_name)
            return

        if self._looks_like_action_command(raw_text):
            print("[ACTION] No command matched. Run 'office help' for office commands.")
